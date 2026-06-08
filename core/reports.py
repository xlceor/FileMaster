import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import os
from core.scanner import scan_folder
from utils.logger import logger

def _style_sheet(ws, theme_color: str, zebra_color: str):
    """Aplica estilos profesionales a una hoja:
    - Encabezado con color temático (negrita, centrado, fondo de color).
    - Celdas con fuentes e interlineados elegantes.
    - Alternancia de colores en filas (zebra striping).
    - Alineación inteligente (ID centrado, textos a la izquierda).
    - Ajuste automático de ancho de columna.
    """
    # Gridlines
    ws.views.sheetView[0].showGridLines = True
    
    # Borders
    thin_side = Side(style="thin", color="D9D9D9")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Header styling
    header_fill = PatternFill("solid", start_color=theme_color.lstrip("#"))
    header_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data row styling
    cell_font = Font(name="Segoe UI", size=10)
    zebra_fill = PatternFill("solid", start_color=zebra_color.lstrip("#"))
    white_fill = PatternFill("solid", start_color="FFFFFF")
    
    # Height of header
    ws.row_dimensions[1].height = 28
    
    # Format Headers
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = cell_border
        
    # Format Data Rows
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        row_fill = zebra_fill if row_idx % 2 == 0 else white_fill
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.font = cell_font
            cell.border = cell_border
            
            # Align primary column (usually index/placas) in center
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # Auto-adjust column widths with safety padding
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def _write_summary_sheet(ws, total: int, found: int, missing: int, extra: int) -> None:
    """Construye una pestaña de Resumen estilizada como un panel ejecutivo."""
    ws.views.sheetView[0].showGridLines = True
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    
    # Header Styling
    header_fill = PatternFill("solid", start_color="5E35B1")
    header_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    
    thin_side = Side(style="thin", color="D9D9D9")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Set headers
    ws.row_dimensions[1].height = 28
    ws["A1"] = "Métrica"
    ws["B1"] = "Valor"
    ws["A1"].fill = header_fill
    ws["B1"].fill = header_fill
    ws["A1"].font = header_font
    ws["B1"].font = header_font
    ws["A1"].alignment = header_align
    ws["B1"].alignment = header_align
    ws["A1"].border = cell_border
    ws["B1"].border = cell_border
    
    metrics = [
        ("Esperados",    total,         "1A237E"), # Navy
        ("Encontrados",  found,         "2E7D32"), # Green
        ("Faltantes",    missing,       "C62828"), # Red
        ("Sobrantes",    extra,         "E65100"), # Orange
        ("% Completitud",
         f"{found/total*100:.1f}%" if total > 0 else "N/A",
         "5E35B1"), # Accent
    ]
    
    for i, (label, val, color) in enumerate(metrics, start=2):
        ws.row_dimensions[i].height = 24
        lc = ws.cell(row=i, column=1, value=label)
        vc = ws.cell(row=i, column=2, value=val)
        
        lc.border = cell_border
        vc.border = cell_border
        lc.alignment = Alignment(horizontal="left", vertical="center")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        
        lc.font = Font(name="Segoe UI", bold=True, size=10, color="212529")
        vc.font = Font(name="Segoe UI", bold=True, size=11, color=color)
        
        # Zebra striping on metrics table
        bg_color = "F8F9FA" if i % 2 == 0 else "FFFFFF"
        lc.fill = PatternFill("solid", start_color=bg_color)
        vc.fill = PatternFill("solid", start_color=bg_color)


def export_names_report(folder: str, recursive: bool, ignore_ext: bool, preprocess: bool, timestamp: str, output_path: str = None) -> str:
    """Genera un archivo Excel que lista los nombres de archivos escaneados."""
    logger.info(f"Exportando nombres de carpeta: {folder}")
    files = scan_folder(folder, recursive, ignore_ext, preprocess)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Archivos"
    
    ws.cell(row=1, column=1, value="Archivo")
    for i, name in enumerate(files, start=2):
        ws.cell(row=i, column=1, value=name)
        
    _style_sheet(ws, "5E35B1", "F3E5F5")
    
    out = output_path if output_path else f"Reporte_Nombres_{timestamp}.xlsx"
    try:
        wb.save(out)
        logger.info(f"Reporte de nombres guardado en {out}")
    except Exception as e:
        logger.error(f"No se pudo guardar el reporte de nombres en {out}: {str(e)}", exc_info=True)
        raise e
        
    return out


def run_comparison(
    source_path: str, master_path: str,
    recursive: bool, ignore_ext: bool, ignore_case: bool, preprocess: bool,
    is_excel_source: bool = False
) -> tuple[dict, dict, dict, dict]:
    """Ejecuta la comparación física/lógica y devuelve los registros clasificados."""
    from core.excel_loader import load_excel
    from core.comparator import compare_dicts, compare_files
    
    logger.info(f"Iniciando run_comparison. Origen: {source_path}, Maestro: {master_path}")
    
    # Load master database
    master_dict = load_excel(master_path, ignore_ext, ignore_case, preprocess)
    if not master_dict:
        logger.warning("El archivo maestro no contiene registros o falló la carga.")
        return {}, {}, {}, {}
        
    # Get master columns template (to align schemas of all sheets)
    first_key = next(iter(master_dict))
    master_cols_template = {col: "" for col in master_dict[first_key].keys()}
    logger.debug(f"Plantilla de columnas del maestro: {list(master_cols_template.keys())}")
    
    if is_excel_source:
        source_dict = load_excel(source_path, ignore_ext, ignore_case, preprocess)
        if not source_dict:
            logger.warning("El archivo origen no contiene registros o falló la carga.")
            return master_dict, {}, {}, {}
            
        found_res, missing_res, extra_res = compare_dicts(master_dict, source_dict)
        
        found_dict = {}
        for id in found_res:
            row = master_dict[id].copy()
            row["Resultado"] = "Encontrado"
            found_dict[id] = row
            
        missing_dict = {}
        for id in missing_res:
            row = master_dict[id].copy()
            row["Resultado"] = "Faltante"
            missing_dict[id] = row
            
        extra_dict = {}
        for id in extra_res:
            row = master_cols_template.copy()
            # Overlay source properties if they match the keys, or write standard values
            row.update(source_dict[id])
            row["Resultado"] = "Sobrante"
            extra_dict[id] = row
            
    else:
        found_list = scan_folder(source_path, recursive, ignore_ext, preprocess)
        if ignore_case:
            found_list = [f.lower() for f in found_list]
            
        expected_ids = set(master_dict.keys())
        results = compare_files(expected_ids, found_list, ignore_case)
        
        found_dict = {}
        for id in results["found"]:
            row = master_dict[id].copy()
            row["Resultado"] = "Encontrado"
            found_dict[id] = row
            
        missing_dict = {}
        for id in results["missing"]:
            row = master_dict[id].copy()
            row["Resultado"] = "Faltante"
            missing_dict[id] = row
            
        extra_dict = {}
        for f in results["extra"]:
            row = master_cols_template.copy()
            row["Resultado"] = "Sobrante (Archivo extra en carpeta)"
            extra_dict[f] = row

    logger.info(f"run_comparison completado. Encontrados: {len(found_dict)}, Faltantes: {len(missing_dict)}, Sobrantes: {len(extra_dict)}")
    return master_dict, found_dict, missing_dict, extra_dict


def report_excel(
    output_path: str, found: dict, missing: dict, extra: dict, master_count: int = None, template_path: str = None
) -> None:
    """Convierte los resultados a DataFrames, los escribe en Excel y aplica estilos."""
    logger.info(f"Generando reporte Excel estructurado y diseñado en: {output_path}")
    
    def prepare_df(data, index_name="placas"):
        if not data:
            return pd.DataFrame(columns=[index_name])
        df = pd.DataFrame.from_dict(data, orient="index")
        if index_name in df.columns:
            df = df.drop(columns=[index_name])
        return df.rename_axis(index_name).reset_index()

    df_enc = prepare_df(found)
    df_fal = prepare_df(missing)
    df_sob = prepare_df(extra)

    try:
        if template_path and os.path.exists(template_path):
            logger.info(f"Usando plantilla personalizada: {template_path}")
            wb = openpyxl.load_workbook(template_path)
        else:
            wb = openpyxl.Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        # Write to sheets
        # If template sheets exist, overwrite them; otherwise, create them
        for name, df in [("Encontrados", df_enc), ("Faltantes", df_fal), ("Sobrantes", df_sob)]:
            if name in wb.sheetnames:
                ws = wb[name]
                # Clear content
                ws.delete_rows(1, ws.max_row)
                # Write header
                for col_num, col_name in enumerate(df.columns, 1):
                    ws.cell(row=1, column=col_num, value=col_name)
                # Write data
                for row_num, row_data in enumerate(df.values, 2):
                    for col_num, val in enumerate(row_data, 1):
                        ws.cell(row=row_num, column=col_num, value=val)
            else:
                ws = wb.create_sheet(name)
                df.to_excel(writer := pd.ExcelWriter(output_path, engine="openpyxl", mode="a"), sheet_name=name, index=False)
                # Note: pd.ExcelWriter doesn't support 'a' mode directly this way in conjunction with loading, 
                # but direct openpyxl manipulation is safer. Let's stick to openpyxl for template data injection.
        
        # Simplified writing using standard openpyxl (safer for templates)
        for name, df in [("Encontrados", df_enc), ("Faltantes", df_fal), ("Sobrantes", df_sob)]:
            ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
            # Clear and write content manually or use pandas with engine='openpyxl'
            # Let's write directly using openpyxl for precision
            ws.delete_rows(1, ws.max_row)
            for c_idx, col in enumerate(df.columns, 1):
                ws.cell(row=1, column=c_idx, value=col)
            for r_idx, row in enumerate(df.values, 2):
                for c_idx, val in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=val)

        # Apply styles (Only if not already styled, or override if preferred)
        _style_sheet(wb["Encontrados"], "2E7D32", "F1F8E9") 
        _style_sheet(wb["Faltantes"], "C62828", "FFEBEE")
        _style_sheet(wb["Sobrantes"], "E65100", "FFF3E0")
        
        # Summary
        if master_count is not None:
            if "Resumen" in wb.sheetnames:
                ws_res = wb["Resumen"]
                ws_res.delete_rows(1, ws_res.max_row)
            else:
                ws_res = wb.create_sheet("Resumen", 0)
            _write_summary_sheet(ws_res, master_count, len(found), len(missing), len(extra))
            
        wb.save(output_path)
        logger.info("Reporte Excel exitoso.")
                
    except Exception as e:
        logger.error(f"Error en reporte: {str(e)}", exc_info=True)
        # Fallback to plain export
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_enc.to_excel(writer, sheet_name="Encontrados", index=False)
            df_fal.to_excel(writer, sheet_name="Faltantes", index=False)
            df_sob.to_excel(writer, sheet_name="Sobrantes", index=False)


def comparison_report(
    source_path: str, master_path: str,
    recursive: bool, ignore_ext: bool, ignore_case: bool, preprocess: bool,
    timestamp: str, is_excel_source: bool = False,
    output_path: str = None, template_path: str = None
) -> str:
    """Orquesta la comparación completa y guarda el reporte final en un paso optimizado."""
    master_dict, found_dict, missing_dict, extra_dict = run_comparison(
        source_path, master_path, recursive, ignore_ext, ignore_case, preprocess, is_excel_source
    )

    out = output_path if output_path else f"Reporte_Comparacion_{timestamp}.xlsx"
    
    # Save sheets AND Summary sheet
    report_excel(out, found_dict, missing_dict, extra_dict, master_count=len(master_dict), template_path=template_path)
    
    return out
