
import os
import openpyxl
from pathlib import Path
from core.comparator import preprocess_placas
import pandas as pd

def load_exel(path: str, ignore_ext: bool = False, preprocess: bool = False):
    if not path or not os.path.exists(path):
        return []

    names = pd.read_excel(path)
    row_name = "placas"

    if row_name in names.columns:
        column_id = row_name
    else:
        column_id = names.columns[0]


    if ignore_ext:
        names['clean_id'] = names[column_id].astype(str).str.remplace(r'\.[^.]+$', '', regex=True)
    if preprocess:
        names['clean_id'] = names[column_id].astype(str).str.split('_').str[0]
        names['clean_id'] = names[column_id].astype(str).str.split('-').str[0]

    dictionary = names.set_index('clean_id').to_dict(orient='index')

    return dictionary


def load_names_from_excel(path: str, ignore_ext: bool = False, preprocess: bool = False) -> list[str]:
    """Lee la primera columna de un Excel y devuelve una lista de nombres."""
    if not path or not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    names = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[0]
        if val is None:
            continue
        name = str(val).strip()
        if ignore_ext:
            name = Path(name).stem
        if preprocess:
            name = preprocess_placas(name)
        if name:
            names.append(name)
    wb.close()
    return sorted(list(set(names)))


def load_master_excel(path: str, ignore_ext: bool = False, ignore_case: bool = False, preprocess: bool = False) -> set[str]:
    """Lee la primera columna del Excel maestro y devuelve un set de nombres."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[0]
        if val is None:
            continue
        name = str(val).strip()
        if ignore_ext:
            name = Path(name).stem
        if preprocess:
            name = preprocess_placas(name)
        if ignore_case:
            name = name.lower()
        if name:
            names.add(name)
    wb.close()
    return names
