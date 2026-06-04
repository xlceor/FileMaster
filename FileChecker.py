
import os
import json
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CONFIG_FILE = "filechecker_config.json"
C = {
    "bg":       "#F8F9FA",
    "surface":  "#FFFFFF",
    "panel":    "#E9ECEF",
    "accent":   "#5E35B1",
    "accent2":  "#7E57C2",
    "green":    "#2E7D32",
    "red":      "#C62828",
    "yellow":   "#F9A825",
    "cyan":     "#00838F",
    "text":     "#212529",
    "subtext":  "#6C757D",
    "border":   "#DEE2E6",
}

def preprocess_placas(name: str) -> str:
    """Extrae los caracteres antes del primer '-' o '_'."""
    import re
    return re.split(r'[-_]', name)[0].strip()


def scan_folder(folder: str, recursive: bool = False, ignore_ext: bool = False, preprocess: bool = False) -> list[str]:
    p = Path(folder)
    pattern = "**/*" if recursive else "*"
    files = []
    for f in p.glob(pattern):
        if f.is_file():
            name = f.stem if ignore_ext else f.name
            name = name.strip()
            if preprocess:
                name = preprocess_placas(name)
            if name:
                files.append(name)
    return sorted(list(set(files)))


def load_names_from_excel(path: str, ignore_ext: bool = False, preprocess: bool = False) -> list[str]:
    """Lee la primera columna de un Excel y devuelve una lista de nombres."""
    if not path or not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    names = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[0]
        if val is None:
            continue
        name = str(val).strip()
        if ignore_ext:
            name = Path(name).stem
        if preprocess:
            name = preprocess_placas(name)
        if name:
            names.append(name)
    wb.close()
    return sorted(list(set(names)))


def load_master_excel(path: str, ignore_ext: bool = False, ignore_case: bool = False, preprocess: bool = False) -> set[str]:
    """Lee la primera columna del Excel maestro y devuelve un set de nombres."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[0]
        if val is None:
            continue
        name = str(val).strip()
        if ignore_ext:
            name = Path(name).stem
        if preprocess:
            name = preprocess_placas(name)
        if ignore_case:
            name = name.lower()
        if name:
            names.add(name)
    wb.close()
    return names


def compare_files(expected: set, found_list: list, ignore_case: bool = False) -> dict:
    """Compara sets y devuelve encontrados, faltantes y sobrantes."""
    found = set(f.lower() if ignore_case else f for f in found_list)
    exp   = set(e.lower() if ignore_case else e for e in expected)
    return {
        "found":   exp & found,
        "missing": exp - found,
        "extra":   found - exp,
    }


def _style_header(cell, bg="#5E35B1", fg="FFFFFF"):
    cell.font      = Font(name="Arial", bold=True, color=fg, size=11)
    cell.fill      = PatternFill("solid", start_color=bg.lstrip("#"))
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_cell(cell, color=None):
    thin = Side(style="thin", color="E0E0E0")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.font = Font(name="Arial", size=10)
    if color:
        cell.font = Font(name="Arial", size=10, color=color.lstrip("#"))


def _col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def export_names_report(folder: str, recursive: bool, ignore_ext: bool, preprocess: bool, timestamp: str) -> str:
    """Genera Reporte_Nombres_<timestamp>.xlsx con los archivos de la carpeta."""
    files = scan_folder(folder, recursive, ignore_ext, preprocess)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Archivos"
    ws.row_dimensions[1].height = 22

    _style_header(ws["A1"], bg="5E35B1")
    ws["A1"] = "Archivo"
    _col_width(ws, 1, 40)

    for i, name in enumerate(files, start=2):
        ws.cell(row=i, column=1, value=name)
        _style_cell(ws.cell(row=i, column=1))
        if i % 2 == 0:
            ws.cell(row=i, column=1).fill = PatternFill("solid", start_color="F0F0F8")

    out = f"Reporte_Nombres_{timestamp}.xlsx"
    wb.save(out)
    return out


def comparison_report(
    source_path: str, master_path: str,
    recursive: bool, ignore_ext: bool, ignore_case: bool, preprocess: bool,
    timestamp: str, is_excel_source: bool = False
) -> str:
    """Genera Reporte_Comparacion_<timestamp>.xlsx con resultados y resumen."""
    expected = load_master_excel(master_path, ignore_ext, ignore_case, preprocess)
    
    if is_excel_source:
        found_list = load_names_from_excel(source_path, ignore_ext, preprocess)
    else:
        found_list = scan_folder(source_path, recursive, ignore_ext, preprocess)
    
    if ignore_case:
        found_list = [f.lower() for f in found_list]

    result = compare_files(expected, found_list, ignore_case)

    found_sorted   = sorted(result["found"])
    missing_sorted = sorted(result["missing"])
    extra_sorted   = sorted(result["extra"])
    max_rows       = max(len(found_sorted), len(missing_sorted), len(extra_sorted), 1)

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Resultados"
    ws1.row_dimensions[1].height = 24

    headers = ["✔ Encontrados", "✖ Faltantes", "⚠ Sobrantes"]
    colors  = ["2E7D32", "C62828", "E65100"]
    widths  = [35, 35, 35]

    for col, (h, c, w) in enumerate(zip(headers, colors, widths), start=1):
        cell = ws1.cell(row=1, column=col, value=h)
        _style_header(cell, bg=c)
        _col_width(ws1, col, w)

    lists = [found_sorted, missing_sorted, extra_sorted]
    txt_c = ["1B5E20", "B71C1C", "BF360C"]

    for row in range(max_rows):
        ws1.row_dimensions[row + 2].height = 18
        for col, (lst, tc) in enumerate(zip(lists, txt_c), start=1):
            val  = lst[row] if row < len(lst) else ""
            cell = ws1.cell(row=row + 2, column=col, value=val)
            _style_cell(cell, color=tc if val else None)
            bg = "F9FBE7" if col == 1 else ("FFEBEE" if col == 2 else "FFF3E0")
            if row % 2 == 0:
                cell.fill = PatternFill("solid", start_color=bg)

    ws2 = wb.create_sheet("Resumen")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 18

    _style_header(ws2["A1"], bg="5E35B1")
    _style_header(ws2["B1"], bg="5E35B1")
    ws2["A1"] = "Métrica"
    ws2["B1"] = "Valor"

    metrics = [
        ("Esperados",    len(expected),         "1A237E"),
        ("Encontrados",  len(found_sorted),      "1B5E20"),
        ("Faltantes",    len(missing_sorted),    "B71C1C"),
        ("Sobrantes",    len(extra_sorted),      "BF360C"),
        ("% Completitud",
         f"{len(found_sorted)/len(expected)*100:.1f}%" if expected else "N/A",
         "5E35B1"),
    ]

    for i, (label, val, color) in enumerate(metrics, start=2):
        ws2.row_dimensions[i].height = 20
        lc = ws2.cell(row=i, column=1, value=label)
        vc = ws2.cell(row=i, column=2, value=val)
        _style_cell(lc)
        _style_cell(vc, color=color)
        lc.font = Font(name="Arial", bold=True, size=10)

    out = f"Reporte_Comparacion_{timestamp}.xlsx"
    wb.save(out)
    return out



def load_config() -> dict:
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_config(data: dict):
    try:
        with open(CONFIG_FILE, "w") as f:
            json.dump(data, f, indent=2)
    except Exception:
        pass



class FileCheckerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("FileChecker v1.1")
        self.resizable(False, False)
        self.configure(bg=C["bg"])

        self.cfg = load_config()

        self.mode               = tk.StringVar(value=self.cfg.get("mode", "export"))
        self.folder_var         = tk.StringVar(value=self.cfg.get("last_folder", ""))
        self.source_excel_var   = tk.StringVar(value=self.cfg.get("last_source_excel", ""))
        self.excel_var          = tk.StringVar(value=self.cfg.get("last_excel", ""))
        self.recursive          = tk.BooleanVar(value=self.cfg.get("recursive", False))
        self.ignore_ext         = tk.BooleanVar(value=self.cfg.get("ignore_ext", False))
        self.ignore_case        = tk.BooleanVar(value=self.cfg.get("ignore_case", True))
        self.status_var         = tk.StringVar(value="Listo")

        self._build_ui()
        self._center_window()


    def _build_ui(self):
        hdr = tk.Frame(self, bg=C["accent"], pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="📁  FileChecker", font=("Arial", 17, "bold"),
                 bg=C["accent"], fg="#FFFFFF").pack()
        tk.Label(hdr, text="Verificador de archivos contra lista maestra",
                 font=("Arial", 9), bg=C["accent"], fg="#E0E0E0").pack()

        body = tk.Frame(self, bg=C["bg"], padx=20, pady=14)
        body.pack(fill="both")

        self.notebook = ttk.Notebook(body)
        self.notebook.pack(fill="x", pady=(0, 10))
        
        self.tab_std = tk.Frame(self.notebook, bg=C["bg"])
        self.tab_plc = tk.Frame(self.notebook, bg=C["bg"])
        
        self.notebook.add(self.tab_std, text="  ESTÁNDAR  ")
        self.notebook.add(self.tab_plc, text="  PLACAS  ")
        
        if self.cfg.get("tab") == "placas":
            self.notebook.select(1)

        self._section(body, "OPERACIÓN")
        mode_f = tk.Frame(body, bg=C["surface"], pady=8, padx=12)
        mode_f.pack(fill="x", pady=(0, 10))

        modes = [
            ("📤  Exportar nombres de carpeta", "export"),
            ("🔍  Comparar carpeta vs Excel maestro", "compare"),
            ("📊  Comparar Excel vs Excel maestro", "compare_excel")
        ]
        for text, val in modes:
            tk.Radiobutton(mode_f, text=text, variable=self.mode, value=val,
                           command=self._toggle_excel_field,
                           bg=C["surface"], fg=C["text"], selectcolor=C["panel"],
                           activebackground=C["surface"], activeforeground=C["text"],
                           font=("Arial", 10)).pack(anchor="w", pady=2)

        self.folder_section = self._section(body, "CARPETA DE ARCHIVOS")
        self.folder_frame = self._path_row(body, self.folder_var, self._browse_folder)

        self.source_excel_section = self._section(body, "EXCEL DE ENTRADA")
        self.source_excel_frame = self._path_row(body, self.source_excel_var, self._browse_source_excel, placeholder="Selecciona el Excel de entrada...")

        self.excel_section_label = self._section(body, "EXCEL MAESTRO")
        self.excel_frame = self._path_row(body, self.excel_var, self._browse_excel, placeholder="Selecciona el archivo Excel maestro...")

        self._section(body, "OPCIONES")
        opts = tk.Frame(body, bg=C["surface"], padx=12, pady=8)
        opts.pack(fill="x", pady=(0, 10))

        for text, var in [
            ("🔁  Búsqueda recursiva (subcarpetas)", self.recursive),
            ("🔤  Ignorar extensión al comparar", self.ignore_ext),
            ("🔡  Ignorar mayúsculas/minúsculas", self.ignore_case),
        ]:
            tk.Checkbutton(opts, text=text, variable=var,
                           bg=C["surface"], fg=C["text"], selectcolor=C["panel"],
                           activebackground=C["surface"], activeforeground=C["text"],
                           font=("Arial", 10)).pack(anchor="w", pady=2)

        btn = tk.Button(body, text="⚡  GENERAR REPORTE",
                        command=self._run,
                        bg=C["accent"], fg="#FFFFFF",
                        font=("Arial", 12, "bold"),
                        relief="flat", cursor="hand2",
                        pady=10, padx=20)
        btn.pack(fill="x", pady=(4, 10))
        btn.bind("<Enter>", lambda e: btn.configure(bg=C["accent2"]))
        btn.bind("<Leave>", lambda e: btn.configure(bg=C["accent"]))

        self.progress = ttk.Progressbar(body, mode="indeterminate", length=400)
        self.progress.pack(fill="x", pady=(0, 6))

        status_f = tk.Frame(body, bg=C["panel"], pady=6, padx=10)
        status_f.pack(fill="x")
        tk.Label(status_f, textvariable=self.status_var,
                 bg=C["panel"], fg=C["cyan"],
                 font=("Arial", 9, "italic")).pack(anchor="w")

        self._toggle_excel_field()

    def _section(self, parent, text):
        lbl = tk.Label(parent, text=text, bg=C["bg"], fg=C["subtext"],
                       font=("Arial", 8, "bold"))
        lbl.pack(anchor="w", pady=(6, 2))
        return lbl

    def _path_row(self, parent, var, cmd, placeholder="Selecciona carpeta..."):
        f = tk.Frame(parent, bg=C["bg"])
        f.pack(fill="x", pady=(0, 4))
        entry = tk.Entry(f, textvariable=var, width=44,
                         bg=C["panel"], fg=C["text"], insertbackground=C["text"],
                         relief="flat", font=("Arial", 9), bd=4)
        entry.pack(side="left", fill="x", expand=True, padx=(0, 6))
        btn = tk.Button(f, text="Examinar…", command=cmd,
                        bg=C["surface"], fg=C["text"],
                        relief="flat", cursor="hand2",
                        font=("Arial", 9), padx=8)
        btn.pack(side="right")
        btn.bind("<Enter>", lambda e: btn.configure(bg=C["accent"], fg="#FFFFFF"))
        btn.bind("<Leave>", lambda e: btn.configure(bg=C["surface"], fg=C["text"]))
        return f


    def _toggle_excel_field(self):
        m = self.mode.get()
        # folder_row is for modes 'export' and 'compare'
        # source_excel_row is for mode 'compare_excel'
        # excel_frame (master) is for 'compare' and 'compare_excel'

        is_export = (m == "export")
        is_compare_folder = (m == "compare")
        is_compare_excel = (m == "compare_excel")

        folder_state = "normal" if (is_export or is_compare_folder) else "disabled"
        self._set_state(self.folder_frame, folder_state)
        self.folder_section.configure(fg=C["subtext"] if folder_state == "normal" else C["border"])

   
        source_state = "normal" if is_compare_excel else "disabled"
        self._set_state(self.source_excel_frame, source_state)
        self.source_excel_section.configure(fg=C["subtext"] if source_state == "normal" else C["border"])

        master_state = "normal" if (is_compare_folder or is_compare_excel) else "disabled"
        self._set_state(self.excel_frame, master_state)
        self.excel_section_label.configure(fg=C["subtext"] if master_state == "normal" else C["border"])

    def _set_state(self, frame, state):
        for w in frame.winfo_children():
            try:
                w.configure(state=state)
                for child in w.winfo_children():
                    child.configure(state=state)
            except Exception:
                pass

    def _browse_folder(self):
        d = filedialog.askdirectory(title="Seleccionar carpeta")
        if d:
            self.folder_var.set(d)

    def _browse_source_excel(self):
        f = filedialog.askopenfilename(
            title="Seleccionar Excel de entrada",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls"), ("Todos", "*.*")]
        )
        if f:
            self.source_excel_var.set(f)

    def _browse_excel(self):
        f = filedialog.askopenfilename(
            title="Seleccionar Excel maestro",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls"), ("Todos", "*.*")]
        )
        if f:
            self.excel_var.set(f)

    def _center_window(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        x = (self.winfo_screenwidth()  - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"+{x}+{y}")

    def _run(self):
        mode = self.mode.get()
        is_placas = (self.notebook.index("current") == 1)
        
        # Validación de entradas
        if mode in ["export", "compare"]:
            folder = self.folder_var.get().strip()
            if not folder or not os.path.isdir(folder):
                messagebox.showerror("Error", "Selecciona una carpeta válida.")
                return
        
        if mode == "compare_excel":
            source_path = self.source_excel_var.get().strip()
            if not source_path or not os.path.isfile(source_path):
                messagebox.showerror("Error", "Selecciona un Excel de entrada válido.")
                return
        else:
            source_path = self.folder_var.get().strip()

        if mode in ["compare", "compare_excel"]:
            excel_master = self.excel_var.get().strip()
            if not excel_master or not os.path.isfile(excel_master):
                messagebox.showerror("Error", "Selecciona un Excel maestro válido.")
                return

        self._save_state()
        self.progress.start(10)
        self.status_var.set("Procesando…")
        self.update()

        try:
            ts  = datetime.now().strftime("%Y-%m-%d_%H-%M")
            rec = self.recursive.get()
            iex = self.ignore_ext.get()
            ica = self.ignore_case.get()

            if mode == "export":
                out = export_names_report(source_path, rec, iex, is_placas, ts)
            else:
                out = comparison_report(
                    source_path, excel_master, 
                    rec, iex, ica, is_placas, ts, 
                    is_excel_source=(mode == "compare_excel")
                )

            self.progress.stop()
            self.status_var.set(f"✔ Reporte generado: {out}")
            if messagebox.askyesno("Listo", f"Reporte generado:\n{out}\n\n¿Abrir ahora?"):
                os.startfile(out) if os.name == "nt" else os.system(f'xdg-open "{out}"')

        except Exception as ex:
            self.progress.stop()
            self.status_var.set(f"✖ Error: {ex}")
            messagebox.showerror("Error", str(ex))

    def _save_state(self):
        save_config({
            "mode":              self.mode.get(),
            "last_folder":       self.folder_var.get(),
            "last_source_excel": self.source_excel_var.get(),
            "last_excel":        self.excel_var.get(),
            "recursive":         self.recursive.get(),
            "ignore_ext":        self.ignore_ext.get(),
            "ignore_case":       self.ignore_case.get(),
            "tab":               "placas" if self.notebook.index("current") == 1 else "standard"
        })


if __name__ == "__main__":
    app = FileCheckerApp()
    app.mainloop()